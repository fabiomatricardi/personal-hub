import os
import re
from openai import OpenAI

TWEET_MAX = 280
MAX_INPUT_CHARS = 10000


def parse_file(filepath: str) -> str:
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".md", ".txt"):
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    elif ext == ".pdf":
        return _parse_pdf(filepath)
    elif ext in (".docx", ".doc"):
        return _parse_docx(filepath)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(filepath)
    else:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()


def _parse_pdf(filepath: str) -> str:
    import pymupdf
    doc = pymupdf.open(filepath)
    pages = []
    for page in doc:
        text = page.get_text()
        if text.strip():
            pages.append(text)
    doc.close()
    return "\n\n".join(pages)


def _parse_docx(filepath: str) -> str:
    try:
        import docx
        doc = docx.Document(filepath)
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return f"[Cannot parse .docx — install python-docx]"


def _parse_excel(filepath: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"## Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                line = " | ".join(str(c) if c is not None else "" for c in row)
                if line.strip():
                    lines.append(line)
        return "\n".join(lines)
    except ImportError:
        return f"[Cannot parse .xlsx — install openpyxl]"


def split_into_tweets(text: str) -> list[str]:
    text = text.strip()
    if len(text) <= TWEET_MAX:
        return [text]

    tweets = []

    # First try splitting on double newlines
    paragraphs = re.split(r"\n\n+", text)

    # If that only gives one chunk, try single newlines
    if len(paragraphs) <= 1:
        paragraphs = re.split(r"\n", text)

    current = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if len(current) + len(para) + 2 <= TWEET_MAX:
            current = f"{current}\n{para}" if current else para
        else:
            if current:
                tweets.append(current)
            if len(para) <= TWEET_MAX:
                current = para
            else:
                sentences = re.split(r"(?<=[.!?])\s+", para)
                current = ""
                for sentence in sentences:
                    if len(current) + len(sentence) + 1 <= TWEET_MAX:
                        current = f"{current} {sentence}" if current else sentence
                    else:
                        if current:
                            tweets.append(current)
                        current = sentence[:TWEET_MAX]
                        if len(sentence) > TWEET_MAX:
                            tweets.append(current)
                            current = ""
                if current:
                    tweets.append(current)
                    current = ""

    if current:
        tweets.append(current)

    return tweets


def _get_provider(config: dict, provider_index: int | None = None) -> dict:
    from backend.services.config import get_active_provider
    if provider_index is not None:
        providers = config.get("providers", [])
        if 0 <= provider_index < len(providers):
            return providers[provider_index]
    return get_active_provider(config)


def _get_model(config: dict, provider_index: int | None = None, model: str | None = None) -> str:
    provider = _get_provider(config, provider_index)
    if model:
        return model
    models = provider.get("models", [])
    return models[0] if models else "gpt-4"


def _get_client(config: dict, provider_index: int | None = None) -> OpenAI:
    provider = _get_provider(config, provider_index)
    api_key = provider.get("api_key", "")
    base_url = provider.get("base_url", "https://api.openai.com/v1").rstrip("/")
    if not api_key:
        raise ValueError("LLM API key not configured. Go to Settings to add it.")
    return OpenAI(api_key=api_key, base_url=base_url)


def call_llm(messages: list[dict], config: dict, provider_index: int | None = None, model: str | None = None) -> str:
    client = _get_client(config, provider_index)
    use_model = _get_model(config, provider_index, model)

    response = client.chat.completions.create(
        model=use_model,
        messages=messages,
        temperature=0.7,
        max_tokens=10000,
    )

    choice = response.choices[0]
    content = choice.message.content

    # Handle reasoning models that may return None content
    if content is None:
        if choice.finish_reason == "length":
            response = client.chat.completions.create(
                model=use_model,
                messages=messages,
                temperature=0.7,
                max_tokens=20000,
            )
            content = response.choices[0].message.content or ""
        else:
            content = getattr(choice.message, 'reasoning', None) or ""

    return content


def call_llm_stream(messages: list[dict], config: dict, provider_index: int | None = None, model: str | None = None):
    client = _get_client(config, provider_index)
    use_model = _get_model(config, provider_index, model)

    stream = client.chat.completions.create(
        model=use_model,
        messages=messages,
        temperature=0.7,
        max_tokens=10000,
        stream=True,
    )

    for chunk in stream:
        if chunk.choices and chunk.choices[0].delta.content:
            yield chunk.choices[0].delta.content


def test_provider_connection(config: dict, provider_index: int | None = None, model: str | None = None) -> dict:
    provider = _get_provider(config, provider_index)
    name = provider.get("name", "Provider")
    api_key = provider.get("api_key", "")
    base_url = provider.get("base_url", "").rstrip("/")

    if not api_key:
        return {"ok": False, "provider": name, "error": "No API key configured"}
    if not base_url:
        return {"ok": False, "provider": name, "error": "No base URL configured"}

    try:
        client = _get_client(config, provider_index)
        use_model = _get_model(config, provider_index, model)

        response = client.chat.completions.create(
            model=use_model,
            messages=[{"role": "user", "content": "Say 'OK' in one word."}],
            max_tokens=50,
        )

        choice = response.choices[0]
        reply = choice.message.content

        # Handle reasoning models returning None
        if reply is None:
            reply = getattr(choice.message, 'reasoning', None) or "OK (reasoning model)"

        return {"ok": True, "provider": name, "model": use_model, "response": reply.strip()}
    except Exception as e:
        return {"ok": False, "provider": name, "error": str(e)}
